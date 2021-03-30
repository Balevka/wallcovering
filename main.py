import datetime
import sys
import time
import requests
from bs4 import BeautifulSoup as bs
import re
import openpyxl
import schedule

start_page = "https://www.mosoboi.ru/oboi/bumazhnye/"
base_url = "https://www.mosoboi.ru"
headers = {'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
           'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.141 Safari/537.36 OPR/73.0.3856.415 (Edition Yx GX)'}

wallcovering_links = []
wallcoverings = []

# Поиск ссылок в каталоге

def get_links(start_page, headers):
    session = requests.Session()

    try:
        request = session.get(start_page, headers=headers)
        request.raise_for_status()
        if request.status_code == 200:
            print('Status 200')
        soup = bs(request.content, 'html.parser')
        links = soup.find_all('div', {'class': 'product-item'})
        for item in links:
            link = str(base_url + item['data-product-click'])
            if link not in wallcovering_links:
                wallcovering_links.append(link)

    except requests.exceptions.HTTPError as err:
        raise SystemExit(err)


# Получение данных

def get_data(links, headers):
    name = ''
    price = ''
    image_link = ''
    companions = {}
    other_colors = {}
    props = {}

    # Список названий характеристик
    properties_list = []

    # Список значений характеристик
    values_list = []

    session = requests.Session()
    for link in links:
        try:
            request = session.get(link, headers=headers)
            request.raise_for_status()

            # Парсинг

            soup = bs(request.content, 'html.parser')

            # Изображение
            block = soup.find('div', {'class': 'detail-image'})
            img = block.findChild('img', {'itemprop': 'image'})
            image = img['src']
            image_link = base_url + image

            # Название

            item_name = soup.find('h1', {'id': 'h1-pages'}).text
            name = re.sub(r'\s+', ' ', item_name)

            # Цена

            block = soup.find('div', {'class': 'today-price'})
            price = block.findChild('span').text

            # Харакатеристики

            specifications = soup.find('div', {'class': 'properties tab-info js-tab-info active'})

            properties = specifications.findChildren('div', {'class': 'tab-content-left'})

            for prop in properties:
                text = prop.text
                if text not in properties_list:
                    properties_list.append(text)

            # Значение характеристик

            values = specifications.findChildren('span', {'class': 'normal'})
            for value in values:
                text = re.sub(r'\s+', ' ', value.text)
                text = text.split('?')[0]
                if text not in values_list:
                    values_list.append(text)

            # Объединение характеристик в один словарь

            if len(values_list) != len(properties_list):
                print('Ошибка парсинга, разное количество характеристик и их значений!')
                props.update({'ошибка': 'Возможна ошибка в парсинге характеристик'})

            for i in range(0, len(values_list)):
                props.update({f'{properties_list[i]}': f'{values_list[i]}'})

            # другие цвета

            try:
                carousel = soup.find('div', {'class': 'head-left pull-left'})
                blocks = carousel.findChildren('div', {'class': 'block interior-carousel-block'})
                temp_dict = {}
                for item in blocks:
                    color_name = item['data-name']
                    color_articul = item['data-articul']
                    dict = {f'{color_name}': f'{color_articul}'}
                    temp_dict.update(dict)
                    other_colors = temp_dict

            except Exception:
                print(Exception)

            # Компаньоны

            try:
                carousel = soup.find('div', {'class': 'head-center pull-left'})
                blocks = carousel.findChildren('div', {'class': 'block interior-carousel-block'})
                temp_dict = {}
                for item in blocks:
                    companion_name = item['data-name']
                    companion_articul = item['data-articul']
                    dict = {f'{companion_name}': f'{companion_articul}'}
                    temp_dict.update(dict)
                    companions = temp_dict

            except Exception:
                print(Exception)

            wallcoverings.append({'image': image_link,
                                  'name': name,
                                  'price': price,
                                  'properties': props,
                                  'companions': companions,
                                  'colors': other_colors})

            values_list = []
            properties_list = []
            # time.sleep(0.3)

        except requests.exceptions.HTTPError as err:
            raise SystemExit(err)


def excel_import(wallcoverings):
    now = datetime.datetime.now()
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        names = ['Изображение', 'Название', 'Цена (руб)', 'Характеристики',
                 'Компаньоны (Название/Артикул)', 'Другие цвета (Название/Артикул)']
        for i in range(1, len(names)+1):
            ws.cell(row=1, column=i).value = names[i-1]
        sheet_row = 2
        for i in wallcoverings:
            image = i['image']
            name = i['name']
            cost = i['price']
            comp = i['companions']
            comp_keys = list(comp.keys())
            comp_values = list(comp.values())
            comp_str = ''
            for j in range(0, len(comp_keys)):
                comp_str = comp_str + comp_keys[j] + ' / ' + comp_values[j] + '\n'

            colors = i['colors']
            colors_keys = list(colors.keys())
            colors_values = list(colors.values())
            colors_str = ''
            for j in range(0, len(colors_keys)):
                colors_str = colors_str + colors_keys[j] + ' / ' + colors_values[j] + '\n'

            prop = i['properties']
            prop_keys = list(prop.keys())
            prop_values = list(prop.values())
            prop_str = ''
            for j in range(0, len(prop_keys)):
                prop_str = prop_str + prop_keys[j] + ' / ' + prop_values[j] + '\n'

            row_list = []
            row_list.append(image)
            row_list.append(name)
            row_list.append(cost)
            row_list.append(prop_str)
            row_list.append(comp_str)
            row_list.append(colors_str)
            print(row_list)
            for j in range(1, len(row_list)+1):
                ws.cell(row=sheet_row, column=j).value = str(row_list[j-1])

            sheet_row+=1
            # print(image)
            # print(name)
            # print(comp_str)
            # print(colors_str)
            # print(prop_str)
            #print(sheet_row)

        date_string = now.strftime('%Y-%m-%d %H-%M-%S')
        wb.save(f'book {date_string}.xlsx')
        wb.close()

    except Exception as e:
        print('Ошибка записи в документ' + str(e))


def main():
    get_links(start_page, headers)
    get_data(wallcovering_links, headers)
    excel_import(wallcoverings)
    sys.exit()


main()
#pprint(wallcoverings)
while True:
    schedule.run_pending()
    time.sleep(1)